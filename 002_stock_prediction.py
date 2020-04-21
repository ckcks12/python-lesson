import os
from openpyxl import Workbook, load_workbook

prices = [100, 110, 120, 140, 180, 230, 300, 390]

for p in prices:
    print(p)

for i in range(1, len(prices)):
    print(prices[i], prices[i] - prices[i-1])

# os.remove('./prices.txt')
# file = open('./prices.txt', 'w+')
# for p in prices:
#     file.writelines(str(p)) # 여기서 str 감싸야하는 오류 한번 잡아주고
# file.close()

# wb = Workbook()
# ws = wb.active
# ws.append(prices)
# wb.save('prices.xlsx')

# wb = load_workbook('prices.xlsx')
# ws = wb.active
# print(ws['B1'].value)

import pygal
chart = pygal.Line()
chart.title = '차트 이름'
chart.x_labels = range(len(prices))
chart.add('price', prices)
chart.render_in_browser()
